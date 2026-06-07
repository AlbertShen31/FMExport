"""Export parsed OCR tables to CSV, Excel, JSON, and Parquet."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.utils import get_column_letter

from fm26_screenshot_exporter.config import OcrToken, ParseResult, ParsedRow


def _rows_to_dataframe(rows: list[ParsedRow], *, normalized: bool = True) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for row in rows:
        source = row.normalized_cells if normalized else row.cells
        record = dict(source)
        record["_row_index"] = row.row_index
        record["_avg_confidence"] = row.avg_confidence
        record["_low_confidence"] = row.low_confidence
        records.append(record)
    if not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


def _ocr_to_dataframe(tokens: list[OcrToken]) -> pd.DataFrame:
    return pd.DataFrame([t.model_dump() for t in tokens])


def _resolve_output(output: str | Path | None, source: Path, suffix: str) -> Path:
    if output is not None:
        out = Path(output)
        if out.suffix.lower() != suffix:
            return out.with_suffix(suffix)
        return out
    return source.with_suffix(suffix)


def escape_csv_cell(value: Any) -> str:
    """Escape a cell value for CSV output (RFC-style quoting)."""
    if value is None:
        return ""
    text = str(value)
    if any(ch in text for ch in [",", '"', "\n", "\r"]):
        return '"' + text.replace('"', '""') + '"'
    return text


def export_csv(
    result: ParseResult,
    output: str | Path | None = None,
    *,
    include_raw: bool = False,
) -> Path:
    source = Path(result.source_path)
    out = _resolve_output(output, source, ".csv")
    df = _rows_to_dataframe(result.rows, normalized=not include_raw)
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out, index=False, encoding="utf-8", quoting=csv.QUOTE_MINIMAL)
    return out


def export_xlsx(
    result: ParseResult,
    output: str | Path | None = None,
) -> Path:
    source = Path(result.source_path)
    out = _resolve_output(output, source, ".xlsx")
    parsed_df = _rows_to_dataframe(result.rows, normalized=True)
    raw_df = _rows_to_dataframe(result.rows, normalized=False)
    ocr_df = _ocr_to_dataframe(result.ocr_tokens)

    metadata_rows = [
        ("source", result.source_path),
        ("profile", result.profile_name),
        ("row_count", len(result.rows)),
        ("column_count", len(result.columns)),
    ]
    for key, value in result.metadata.items():
        metadata_rows.append((str(key), str(value)))
    metadata_df = pd.DataFrame(metadata_rows, columns=["key", "value"])

    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        parsed_df.to_excel(writer, index=False, sheet_name="Parsed Data")
        raw_df.to_excel(writer, index=False, sheet_name="Raw OCR")
        ocr_df.to_excel(writer, index=False, sheet_name="OCR Tokens")
        metadata_df.to_excel(writer, index=False, sheet_name="Metadata")

        for sheet_name in ("Parsed Data", "Raw OCR"):
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions
            df = parsed_df if sheet_name == "Parsed Data" else raw_df
            for idx, column in enumerate(df.columns, start=1):
                letter = get_column_letter(idx)
                lengths = [len(str(column))]
                if column in df.columns:
                    lengths.extend(
                        len(str(v)) if v is not None else 0 for v in df[column].head(50)
                    )
                ws.column_dimensions[letter].width = min(max(lengths) + 2, 50)

    return out


def export_json(
    result: ParseResult,
    output: str | Path | None = None,
    *,
    indent: int = 2,
) -> Path:
    source = Path(result.source_path)
    out = _resolve_output(output, source, ".json")
    payload = {
        "source": result.source_path,
        "profile": result.profile_name,
        "columns": result.columns,
        "records": [
            {
                "row_index": row.row_index,
                "cells": row.cells,
                "normalized_cells": row.normalized_cells,
                "avg_confidence": row.avg_confidence,
                "low_confidence": row.low_confidence,
            }
            for row in result.rows
        ],
        "metadata": result.metadata,
    }
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=indent, default=str), encoding="utf-8")
    return out


def export_parquet(
    result: ParseResult,
    output: str | Path | None = None,
) -> Path:
    try:
        import pyarrow  # noqa: F401
    except ImportError as exc:
        raise ImportError(
            "pyarrow is required for Parquet export. "
            "Install with: pip install 'fm26-screenshot-exporter[parquet]'"
        ) from exc

    source = Path(result.source_path)
    out = _resolve_output(output, source, ".parquet")
    df = _rows_to_dataframe(result.rows, normalized=True)
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(out, index=False)
    return out


def export_by_extension(result: ParseResult, output: str | Path) -> Path:
    out = Path(output)
    suffix = out.suffix.lower()
    if suffix == ".csv":
        return export_csv(result, out)
    if suffix in {".xlsx", ".xls"}:
        return export_xlsx(result, out)
    if suffix == ".json":
        return export_json(result, out)
    if suffix == ".parquet":
        return export_parquet(result, out)
    raise ValueError(f"Unsupported output format: {suffix}")
